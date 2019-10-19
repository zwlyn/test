#-*- encoding:utf_8 -*-
from log import logger

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = load_workbook('test.xlsx')

logger.info(wb.sheetnames)
sheet = wb.get_sheet_by_name(wb.sheetnames[0])

logger.info(sheet['H2'].coordinate) 

logger.info(sheet.cell(row=3, column=1).value)
logger.info(sheet.max_row)
logger.info(sheet.max_column)

sheet['H2'].value = 666666666666666

wb.save("test.xlsx")

#sheet['H2'].coodinaate 输出对应表名 ： H2

#创建 新表
# wb = Workbook()
# sheet = wb.active
# sheet.title = "New Shit"
# sheet['F6'] = "Hello SPA!"
# for i in range(1,10):
# 	sheet["A%d" % i] = i+1
# sheet["E1"].value = "=SUM(A:A)"  #使用excel的内置函数
#wb.save("3.xlsx")

#解决excel的列到了Z之后成了AA,AB的难搞东西：
#sheet.cell(row=1, column=2)

#获得表中的行数和列数,并且使用get_column_letter(500),配合使用访问列表中的每个元素
#sheet.max_row
#sheet.max_column
#get_column_letter(27)
#get_column_letter(900)
#如，满屏666：
# for column in range(1,100):
# 	for row in range(1,100):
# 		columnName = get_column_letter(column)
# 		sheet["%s%d" % (columnName,row)].value = 666
# 		logger.info("%s%d" % (columnName,row))

#设置单元格大小
# >>> sheet['A1'] = 'Tall row'
# >>> sheet['B2'] = 'Wide column'
# >>> sheet.row_dimensions[1].height = 70
# >>> sheet.column_dimensions['B'].width = 20

#合并和拆分单元格子
# >>> import openpyxl
# >>> wb = openpyxl.Workbook()
# >>> sheet = wb.active
# >>> sheet.merge_cells('A1:D3')
# >>> sheet['A1'] = 'Twelve cells merged together.'
# >>> sheet.merge_cells('C5:D5')
#sheet.unmerge_cells('A1:D3')
# >>> sheet['C5'] = 'Two merged cells.'
# >>> wb.save('merged.xlsx')